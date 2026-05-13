from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from api.cache import daily_cache, delete_cached, get_cached, set_cached


def _make_crb_workbook(rows: list[tuple[str, float]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row_idx in range(1, 6):
        sheet.cell(row=row_idx, column=1, value=f"Header {row_idx}")
    for row_idx, (date, value) in enumerate(rows, start=6):
        sheet.cell(row=row_idx, column=1, value=date)
        sheet.cell(row=row_idx, column=2, value=value)
    buf = BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _isolate_crb_store(monkeypatch, tmp_path: Path):
    os.environ["STATE_STORAGE_BACKEND"] = "local"

    from api.routers import economic_growth as router

    monkeypatch.setattr(router, "CRB_LOCAL_PATH", tmp_path / "crb.xlsx")
    monkeypatch.setattr(router, "CRB_METADATA_LOCAL_PATH", tmp_path / "crb.json")
    monkeypatch.setattr(router, "CRB_GCS_KEY", "test/economic_growth/crb.xlsx")
    monkeypatch.setattr(router, "CRB_METADATA_GCS_KEY", "test/economic_growth/crb.json")
    delete_cached(daily_cache, router.ECONOMIC_GROWTH_CACHE_KEY)
    return router


def _stub_market_fetch(monkeypatch):
    import macro.economic_growth.economic_growth as eg

    def fake_fetch_all_returns(_tickers, periods, category_name, crb_returns=None):
        if category_name == "Commodities":
            return {eg.CRB_INDEX_NAME: dict(crb_returns or {period: None for period in periods})}
        return {}

    monkeypatch.setattr(eg, "fetch_all_returns", fake_fetch_all_returns)
    return eg


def test_read_crb_from_xls_accepts_workbook_bytes():
    import macro.economic_growth.economic_growth as eg

    payload = _make_crb_workbook(
        [
            ("2026-03-20", 600.0),
            ("2026-03-27", 615.37),
        ]
    )

    df = eg.read_crb_from_xls(payload, filename="crb.xlsx")

    assert df is not None
    assert len(df) == 2
    assert df["date"].iloc[-1].date().isoformat() == "2026-03-27"
    assert float(df["value"].iloc[-1]) == 615.37


def test_crb_upload_saves_metadata_invalidates_cache_and_get_uses_managed_file(auth_client, monkeypatch, tmp_path):
    router = _isolate_crb_store(monkeypatch, tmp_path)
    _stub_market_fetch(monkeypatch)
    payload = _make_crb_workbook(
        [
            ("2025-03-27", 500.0),
            ("2026-03-27", 615.37),
        ]
    )

    set_cached(daily_cache, router.ECONOMIC_GROWTH_CACHE_KEY, {"sentinel": True})
    upload = auth_client.post(
        "/api/economic-growth/crb-file",
        files={
            "file": (
                "crb.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert upload.status_code == 200
    uploaded = upload.json()["crb"]
    assert uploaded["filename"] == "crb.xlsx"
    assert uploaded["rows"] == 2
    assert uploaded["latest_date"] == "2026-03-27"
    assert uploaded["latest_value"] == 615.37
    assert get_cached(daily_cache, router.ECONOMIC_GROWTH_CACHE_KEY) is None

    response = auth_client.get("/api/economic-growth", params={"force_refresh": "true"})

    assert response.status_code == 200
    data = response.json()
    assert data["crb_available"] is True
    assert data["crb_filename"] == "crb.xlsx"
    assert data["crb_latest_date"] == "2026-03-27"
    assert data["crb_latest_value"] == 615.37
    assert data["crb_rows"] == 2
    assert data["crb_uploaded_at"] == uploaded["uploaded_at"]


def test_crb_upload_rejects_invalid_or_empty_file_without_replacing_existing(auth_client, monkeypatch, tmp_path):
    router = _isolate_crb_store(monkeypatch, tmp_path)
    valid_payload = _make_crb_workbook(
        [
            ("2026-03-20", 600.0),
            ("2026-03-27", 615.37),
        ]
    )
    valid = auth_client.post(
        "/api/economic-growth/crb-file",
        files={
            "file": (
                "crb.xlsx",
                valid_payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert valid.status_code == 200

    invalid = auth_client.post(
        "/api/economic-growth/crb-file",
        files={
            "file": (
                "crb.xlsx",
                b"not an excel workbook",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    empty = auth_client.post(
        "/api/economic-growth/crb-file",
        files={
            "file": (
                "empty.xlsx",
                b"",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert invalid.status_code == 422
    assert empty.status_code == 422
    assert router.CRB_LOCAL_PATH.read_bytes() == valid_payload


def test_crb_upload_rejects_endpoint_oversized_file(auth_client, monkeypatch, tmp_path):
    router = _isolate_crb_store(monkeypatch, tmp_path)
    monkeypatch.setattr(router, "MAX_CRB_UPLOAD_SIZE_BYTES", 4)

    upload = auth_client.post(
        "/api/economic-growth/crb-file",
        files={
            "file": (
                "crb.xlsx",
                b"12345",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert upload.status_code == 413


def test_economic_growth_get_falls_back_to_bundled_crb_when_no_managed_file(auth_client, monkeypatch, tmp_path):
    router = _isolate_crb_store(monkeypatch, tmp_path)
    eg = _stub_market_fetch(monkeypatch)
    fallback = tmp_path / "fallback" / "crb.xlsx"
    fallback.parent.mkdir()
    fallback.write_bytes(
        _make_crb_workbook(
            [
                ("2025-01-03", 540.0),
                ("2026-01-02", 580.5),
            ]
        )
    )
    monkeypatch.setattr(eg, "DEFAULT_CRB_PATH", fallback)
    assert not router.CRB_LOCAL_PATH.exists()

    response = auth_client.get("/api/economic-growth", params={"force_refresh": "true"})

    assert response.status_code == 200
    data = response.json()
    assert data["crb_available"] is True
    assert data["crb_filename"] == "crb.xlsx"
    assert data["crb_uploaded_at"] is None
    assert data["crb_latest_date"] == "2026-01-02"
    assert data["crb_latest_value"] == 580.5
    assert data["crb_rows"] == 2

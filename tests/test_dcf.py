import math
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from api.routers.dcf import DCFValuationRequest
from equities.valuation import dcf


def _mock_yf_data():
    annual_date = pd.Timestamp("2025-12-31")
    income_stmt = pd.DataFrame(
        [[1_000.0]],
        index=["Total Revenue"],
        columns=[annual_date],
    )
    balance_sheet = pd.DataFrame(
        [[100.0], [20.0]],
        index=["Total Debt", "Cash And Cash Equivalents"],
        columns=[annual_date],
    )
    empty = pd.DataFrame()
    return {
        "info": {
            "longName": "Test Co",
            "currentPrice": 10.0,
            "regularMarketPrice": 10.0,
            "sharesOutstanding": 100.0,
        },
        "income_stmt": income_stmt,
        "quarterly_income_stmt": empty,
        "balance_sheet": balance_sheet,
        "quarterly_balance_sheet": empty,
        "cashflow": empty,
        "quarterly_cashflow": empty,
        "prices": empty,
    }


def _assumptions(years: int) -> dict:
    return {
        "revenue_growth_rates": [0.10] * years,
        "ebitda_margin": [0.20 + (i * 0.01) for i in range(years)],
        "tax_rate": [0.21] * years,
        "da_pct_revenue": [0.03] * years,
        "nwc_pct_revenue": [0.05] * years,
        "capex_pct_revenue": [0.04] * years,
        "wacc": 0.10,
        "terminal_growth_rates": {"bear": 0.02, "base": 0.03, "bull": 0.04},
        "exit_ev_ebitda": {"bear": 8.0, "base": 10.0, "bull": 12.0},
        "exit_ev_revenue": {"bear": 2.0, "base": 3.0, "bull": 4.0},
    }


def _historical_payload(ticker: str = "TEST") -> dict:
    return {
        "ticker": ticker,
        "company_name": "Test Co",
        "current_price": 10.0,
        "shares_outstanding": 100.0,
        "net_debt": 80.0,
        "base_revenue": 1_000.0,
        "data_source": "yfinance",
        "ebitda": [{"fiscal_year": "2025", "revenue": 1_000.0, "ebitda": 200.0, "ebitda_margin": 20.0, "avg": 20.0}],
        "depreciation": [{"fiscal_year": "2025", "revenue": 1_000.0, "da": 30.0, "da_pct_rev": 3.0, "avg": 3.0}],
        "capex": [{"fiscal_year": "2025", "revenue": 1_000.0, "capex": 40.0, "capex_pct_rev": 4.0, "avg": 4.0}],
        "nwc": [{"fiscal_year": "2025", "revenue": 1_000.0, "nwc": 50.0, "nwc_pct_rev": 5.0, "avg": 5.0}],
        "ev_ebitda": [{"quarter_end": "2025-12-31", "ev": 1_100.0, "ev_ebitda": 10.0, "avg": 10.0}],
        "rev_multiple": [{"quarter_end": "2025-12-31", "ev": 1_100.0, "ev_revenue": 2.5, "avg": 2.5}],
    }


def test_run_valuation_supports_dynamic_projection_years(monkeypatch):
    monkeypatch.setattr(dcf, "_fetch_yfinance_data", lambda ticker: _mock_yf_data())

    for years in (5, 6, 8):
        result = dcf.run_valuation("TEST", _assumptions(years))

        assert len(result["projection"]) == years
        assert result["projection"][-1]["year"] == years
        base_gordon = result["valuations"]["gordon_growth"]["base"]
        assert math.isclose(
            base_gordon["pv_terminal_value"],
            base_gordon["terminal_value"] / ((1 + 0.10) ** years),
        )


def test_run_valuation_uses_year_specific_assumptions(monkeypatch):
    monkeypatch.setattr(dcf, "_fetch_yfinance_data", lambda ticker: _mock_yf_data())

    assumptions = _assumptions(5)
    assumptions["revenue_growth_rates"] = [0.0] * 5
    assumptions["ebitda_margin"] = [0.20, 0.30, 0.40, 0.50, 0.60]

    result = dcf.run_valuation("TEST", assumptions)

    assert result["projection"][0]["ebitda"] == 200.0
    assert result["projection"][1]["ebitda"] == 300.0
    assert result["projection"][4]["ebitda"] == 600.0


def test_dcf_request_accepts_scalar_assumptions_for_compatibility():
    req = DCFValuationRequest(
        ticker="TEST",
        revenue_growth_rates=[0.10] * 6,
        ebitda_margin=0.25,
        tax_rate=0.21,
        da_pct_revenue=0.03,
        nwc_pct_revenue=0.05,
        capex_pct_revenue=0.04,
        wacc=0.10,
        exit_ev_ebitda={"bear": 8.0, "base": 10.0, "bull": 12.0},
        exit_ev_revenue={"bear": 2.0, "base": 3.0, "bull": 4.0},
    )

    dumped = req.model_dump()
    assert dumped["ebitda_margin"] == [0.25] * 6
    assert dumped["tax_rate"] == [0.21] * 6
    assert dumped["capex_pct_revenue"] == [0.04] * 6


def test_ev_revenue_multiples_do_not_require_ebitda():
    quarter_dates = pd.to_datetime(["2024-03-31", "2024-06-30", "2024-09-30", "2024-12-31", "2025-03-31"])
    quarterly_income = pd.DataFrame(
        [[100.0, 110.0, 120.0, 130.0, 140.0]],
        index=["Total Revenue"],
        columns=quarter_dates,
    )
    quarterly_balance = pd.DataFrame(
        [[10.0] * 5, [2.0] * 5],
        index=["Total Debt", "Cash And Cash Equivalents"],
        columns=quarter_dates,
    )
    prices = pd.DataFrame({"Close": [10.0, 11.0, 12.0, 13.0, 14.0]}, index=quarter_dates)

    ev_ebitda, ev_revenue = dcf._compute_multiples_yf(
        quarterly_income,
        quarterly_balance,
        prices,
        {"sharesOutstanding": 10.0},
    )

    assert ev_ebitda == []
    assert len(ev_revenue) == 2
    assert all(row["ev_revenue"] > 0 for row in ev_revenue)


def test_dcf_historical_route_includes_source_registry(monkeypatch):
    from api.cache import invalidate_all
    from api.routers import dcf as dcf_router

    invalidate_all()
    monkeypatch.setattr(dcf, "get_historical_data", lambda ticker: {"ticker": ticker, "data_source": "edgar"})

    result = dcf_router.get_dcf_historical("TEST")

    assert result["_meta"]["source_registry"]["source_id"] == "dcf_historical"


def test_dcf_excel_download_returns_formula_workbook(auth_client, monkeypatch):
    monkeypatch.setattr(dcf, "_fetch_yfinance_data", lambda ticker: _mock_yf_data())
    monkeypatch.setattr(dcf, "get_historical_data", lambda ticker: _historical_payload(ticker))

    payload = {"ticker": "test", **_assumptions(5)}
    resp = auth_client.post("/api/dcf/valuation/excel", json=payload)

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert 'filename="TEST_dcf_model.xlsx"' in resp.headers["content-disposition"]

    workbook = load_workbook(BytesIO(resp.content), data_only=False)
    assert {"Summary", "Assumptions", "Historical", "Projection", "Valuation", "Checks", "Sources"}.issubset(
        workbook.sheetnames
    )
    assert workbook["Projection"]["B4"].value.startswith("=")
    assert workbook["Projection"]["B18"].value.startswith("=")
    assert workbook["Valuation"]["I10"].value.startswith("=")
    assert workbook["Checks"]["B6"].value.startswith("=")


def test_dcf_excel_download_supports_eight_projection_years(auth_client, monkeypatch):
    monkeypatch.setattr(dcf, "_fetch_yfinance_data", lambda ticker: _mock_yf_data())
    monkeypatch.setattr(dcf, "get_historical_data", lambda ticker: _historical_payload(ticker))

    payload = {"ticker": "TEST", **_assumptions(8)}
    resp = auth_client.post("/api/dcf/valuation/excel", json=payload)

    assert resp.status_code == 200
    workbook = load_workbook(BytesIO(resp.content), data_only=False)
    assert workbook["Projection"]["I3"].value == "Year 8"
    assert workbook["Projection"]["I18"].value.startswith("=")


def test_dcf_excel_download_reuses_request_validation(auth_client):
    payload = {"ticker": "TEST", **_assumptions(5), "wacc": 1.2}

    resp = auth_client.post("/api/dcf/valuation/excel", json=payload)

    assert resp.status_code == 422

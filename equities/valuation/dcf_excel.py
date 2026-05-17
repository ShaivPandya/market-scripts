"""Excel workbook builder for DCF valuation exports."""

from __future__ import annotations

import math
from collections.abc import Mapping, Sequence
from datetime import UTC, datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

SCENARIOS: tuple[str, ...] = ("bear", "base", "bull")
SCENARIO_LABELS = {"bear": "Bear", "base": "Base", "bull": "Bull"}

CURRENCY_FORMAT = "$#,##0;[Red]($#,##0);-"
PRICE_FORMAT = "$0.00;[Red]($0.00);-"
PERCENT_FORMAT = "0.0%;[Red](0.0%);-"
MULTIPLE_FORMAT = "0.0x;[Red](0.0x);-"
NUMBER_FORMAT = "#,##0;[Red](#,##0);-"
DATE_FORMAT = "yyyy-mm-dd"

TITLE_FONT = Font(bold=True, size=16, color="111827")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, color="111827")
LABEL_FONT = Font(bold=True, color="374151")
INPUT_FONT = Font(color="0000FF")
FORMULA_FONT = Font(color="000000")
NOTE_FONT = Font(color="6B7280", italic=True)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="DCFCE7")
CHECK_FILL = PatternFill("solid", fgColor="FEE2E2")
THIN_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))

ASSUMPTION_ROWS = {
    "revenue_growth_rates": 15,
    "ebitda_margin": 16,
    "tax_rate": 17,
    "da_pct_revenue": 18,
    "nwc_pct_revenue": 19,
    "capex_pct_revenue": 20,
}
BASE_REVENUE_CELL = "Assumptions!$B$8"
PROJECTION_YEARS_CELL = "Assumptions!$B$10"
CURRENT_PRICE_CELL = "Assumptions!$B$5"
SHARES_CELL = "Assumptions!$B$6"
NET_DEBT_CELL = "Assumptions!$B$7"
WACC_CELL = "Assumptions!$B$23"
TERMINAL_GROWTH_ROW = 27
EV_EBITDA_ROW = 28
EV_REVENUE_ROW = 29


def build_dcf_workbook_bytes(
    valuation: Mapping[str, Any],
    historical: Mapping[str, Any] | None = None,
) -> bytes:
    """Build a formula-driven DCF workbook and return XLSX bytes."""
    workbook = _build_workbook(valuation, historical or {})
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_workbook(valuation: Mapping[str, Any], historical: Mapping[str, Any]) -> Workbook:
    workbook = Workbook()
    workbook.properties.creator = "Talisman"
    workbook.properties.title = f"{valuation.get('ticker', 'Ticker')} DCF Model"
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    summary = workbook.active
    summary.title = "Summary"
    assumptions = workbook.create_sheet("Assumptions")
    historical_sheet = workbook.create_sheet("Historical")
    projection = workbook.create_sheet("Projection")
    valuation_sheet = workbook.create_sheet("Valuation")
    checks = workbook.create_sheet("Checks")
    sources = workbook.create_sheet("Sources")

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

    projection_years = len(valuation.get("projection") or [])
    _build_assumptions_sheet(assumptions, valuation, projection_years)
    _build_projection_sheet(projection, projection_years)
    valuation_rows = _build_valuation_sheet(valuation_sheet, projection_years)
    _build_summary_sheet(summary, valuation, valuation_rows)
    _build_historical_sheet(historical_sheet, historical)
    _build_checks_sheet(checks, projection_years)
    _build_sources_sheet(sources, valuation, historical)

    return workbook


def _build_assumptions_sheet(sheet: Worksheet, valuation: Mapping[str, Any], projection_years: int) -> None:
    _title(sheet, "DCF Model Assumptions", max(6, projection_years + 1))
    assumptions = _mapping(valuation.get("assumptions_used"))
    ticker = str(valuation.get("ticker") or assumptions.get("ticker") or "").upper()

    metadata = [
        ("Ticker", ticker),
        ("Company Name", valuation.get("company_name")),
        ("Current Price", _number_or_none(valuation.get("current_price")), PRICE_FORMAT),
        ("Shares Outstanding", _number_or_none(valuation.get("shares_outstanding")), NUMBER_FORMAT),
        ("Net Debt", _number_or_none(valuation.get("net_debt")), CURRENCY_FORMAT),
        ("Base Revenue", _number_or_none(valuation.get("base_revenue")), CURRENCY_FORMAT),
        ("Base Year", valuation.get("base_year")),
        ("Projection Years", projection_years, NUMBER_FORMAT),
        ("Valuation As Of", datetime.now(UTC).date(), DATE_FORMAT),
    ]
    for row, item in enumerate(metadata, start=3):
        label, value, *fmt = item
        _label(sheet.cell(row=row, column=1), str(label))
        cell = sheet.cell(row=row, column=2, value=value)
        cell.number_format = fmt[0] if fmt else "General"
        cell.font = INPUT_FONT
        cell.fill = INPUT_FILL

    _section(sheet, 13, "Operating Assumptions")
    _header_row(sheet, 14, ["Metric", *[f"Year {idx}" for idx in range(1, projection_years + 1)]])
    assumption_labels = [
        ("Revenue Growth", "revenue_growth_rates", PERCENT_FORMAT),
        ("EBITDA Margin", "ebitda_margin", PERCENT_FORMAT),
        ("Tax Rate", "tax_rate", PERCENT_FORMAT),
        ("D&A % Revenue", "da_pct_revenue", PERCENT_FORMAT),
        ("NWC % Revenue", "nwc_pct_revenue", PERCENT_FORMAT),
        ("CapEx % Revenue", "capex_pct_revenue", PERCENT_FORMAT),
    ]
    for label, key, number_format in assumption_labels:
        row = ASSUMPTION_ROWS[key]
        _label(sheet.cell(row=row, column=1), label)
        for idx, value in enumerate(_series(assumptions.get(key), projection_years), start=2):
            cell = sheet.cell(row=row, column=idx, value=_number_or_none(value))
            cell.number_format = number_format
            cell.font = INPUT_FONT
            cell.fill = INPUT_FILL

    _section(sheet, 22, "Discount Rate")
    _label(sheet.cell(row=23, column=1), "WACC")
    wacc = sheet.cell(row=23, column=2, value=_number_or_none(assumptions.get("wacc")))
    wacc.number_format = PERCENT_FORMAT
    wacc.font = INPUT_FONT
    wacc.fill = INPUT_FILL

    _section(sheet, 25, "Exit Assumptions")
    _header_row(sheet, 26, ["Method", "Bear", "Base", "Bull"])
    terminal_growth = _mapping(assumptions.get("terminal_growth_rates"))
    exit_ebitda = _mapping(assumptions.get("exit_ev_ebitda"))
    exit_revenue = _mapping(assumptions.get("exit_ev_revenue"))
    _scenario_row(sheet, TERMINAL_GROWTH_ROW, "Terminal Growth", terminal_growth, PERCENT_FORMAT)
    _scenario_row(sheet, EV_EBITDA_ROW, "EV/EBITDA Exit", exit_ebitda, MULTIPLE_FORMAT)
    _scenario_row(sheet, EV_REVENUE_ROW, "EV/Revenue Exit", exit_revenue, MULTIPLE_FORMAT)

    _set_widths(sheet, {"A": 24, "B": 18, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14})
    sheet.freeze_panes = "B15"


def _build_projection_sheet(sheet: Worksheet, projection_years: int) -> None:
    _title(sheet, "DCF Projection", max(7, projection_years + 1))
    _header_row(sheet, 2, ["", *[idx for idx in range(1, projection_years + 1)]])
    _header_row(sheet, 3, ["Metric", *[f"Year {idx}" for idx in range(1, projection_years + 1)]])

    labels = {
        4: "Revenue",
        5: "Revenue Growth",
        6: "EBITDA",
        7: "EBITDA Margin",
        8: "D&A",
        9: "EBIT",
        10: "Tax Rate",
        11: "NOPAT",
        12: "NWC",
        13: "Change in NWC",
        14: "CapEx",
        15: "Unlevered FCF",
        17: "Discount Factor",
        18: "PV of UFCF",
    }
    for row, label in labels.items():
        _label(sheet.cell(row=row, column=1), label)

    for idx in range(projection_years):
        col = 2 + idx
        letter = get_column_letter(col)
        assumption_col = get_column_letter(col)
        previous_letter = get_column_letter(col - 1) if idx else ""

        revenue_growth = f"Assumptions!{assumption_col}${ASSUMPTION_ROWS['revenue_growth_rates']}"
        ebitda_margin = f"Assumptions!{assumption_col}${ASSUMPTION_ROWS['ebitda_margin']}"
        tax_rate = f"Assumptions!{assumption_col}${ASSUMPTION_ROWS['tax_rate']}"
        da_pct = f"Assumptions!{assumption_col}${ASSUMPTION_ROWS['da_pct_revenue']}"
        nwc_pct = f"Assumptions!{assumption_col}${ASSUMPTION_ROWS['nwc_pct_revenue']}"
        capex_pct = f"Assumptions!{assumption_col}${ASSUMPTION_ROWS['capex_pct_revenue']}"

        _formula(
            sheet.cell(row=4, column=col),
            f"={BASE_REVENUE_CELL}*(1+{revenue_growth})" if idx == 0 else f"={previous_letter}4*(1+{revenue_growth})",
            CURRENCY_FORMAT,
        )
        _formula(sheet.cell(row=5, column=col), f"={revenue_growth}", PERCENT_FORMAT)
        _formula(sheet.cell(row=6, column=col), f"={letter}4*{ebitda_margin}", CURRENCY_FORMAT)
        _formula(sheet.cell(row=7, column=col), f"={letter}6/{letter}4", PERCENT_FORMAT)
        _formula(sheet.cell(row=8, column=col), f"={letter}4*{da_pct}", CURRENCY_FORMAT)
        _formula(sheet.cell(row=9, column=col), f"={letter}6-{letter}8", CURRENCY_FORMAT)
        _formula(sheet.cell(row=10, column=col), f"={tax_rate}", PERCENT_FORMAT)
        _formula(sheet.cell(row=11, column=col), f"={letter}9*(1-{letter}10)", CURRENCY_FORMAT)
        _formula(sheet.cell(row=12, column=col), f"={letter}4*{nwc_pct}", CURRENCY_FORMAT)
        delta_formula = f"={letter}12-{BASE_REVENUE_CELL}*{nwc_pct}" if idx == 0 else f"={letter}12-{previous_letter}12"
        _formula(sheet.cell(row=13, column=col), delta_formula, CURRENCY_FORMAT)
        _formula(sheet.cell(row=14, column=col), f"={letter}4*{capex_pct}", CURRENCY_FORMAT)
        _formula(sheet.cell(row=15, column=col), f"={letter}11+{letter}8-{letter}14-{letter}13", CURRENCY_FORMAT)
        _formula(sheet.cell(row=17, column=col), f"=1/(1+{WACC_CELL})^{letter}$2", "0.000x")
        _formula(sheet.cell(row=18, column=col), f"={letter}15*{letter}17", CURRENCY_FORMAT)

    _subtotal_border(sheet, 15, projection_years + 1)
    _subtotal_border(sheet, 18, projection_years + 1)
    _set_widths(sheet, {"A": 24, **{get_column_letter(col): 16 for col in range(2, projection_years + 2)}})
    sheet.freeze_panes = "B4"


def _build_valuation_sheet(sheet: Worksheet, projection_years: int) -> dict[tuple[str, str], int]:
    _title(sheet, "Valuation", 10)
    _label_value_formula(sheet, 3, "Current Price", f"={CURRENT_PRICE_CELL}", PRICE_FORMAT)
    _label_value_formula(sheet, 4, "Shares Outstanding", f"={SHARES_CELL}", NUMBER_FORMAT)
    _label_value_formula(sheet, 5, "Net Debt", f"={NET_DEBT_CELL}", CURRENCY_FORMAT)

    last_projection_col = get_column_letter(projection_years + 1)
    _label_value_formula(sheet, 6, "PV Forecast UFCF", f"=SUM(Projection!B18:{last_projection_col}18)", CURRENCY_FORMAT)

    _header_row(
        sheet,
        8,
        [
            "Method",
            "Scenario",
            "Terminal Assumption",
            "Terminal Value",
            "PV Terminal Value",
            "Enterprise Value",
            "Net Debt",
            "Equity Value",
            "Value/Share",
            "Upside/(Downside)",
        ],
    )

    last_revenue = f"Projection!{last_projection_col}4"
    last_ebitda = f"Projection!{last_projection_col}6"
    last_ufcf = f"Projection!{last_projection_col}15"
    row_map: dict[tuple[str, str], int] = {}
    row = 9
    methods = [
        ("gordon_growth", "Gordon Growth", TERMINAL_GROWTH_ROW, PERCENT_FORMAT, last_ufcf),
        ("ev_ebitda_exit", "EV/EBITDA Exit", EV_EBITDA_ROW, MULTIPLE_FORMAT, last_ebitda),
        ("ev_revenue_exit", "EV/Revenue Exit", EV_REVENUE_ROW, MULTIPLE_FORMAT, last_revenue),
    ]

    for method_key, method_label, assumption_row, assumption_format, terminal_metric_ref in methods:
        for scenario_idx, scenario in enumerate(SCENARIOS, start=2):
            assumption_ref = f"Assumptions!{get_column_letter(scenario_idx)}${assumption_row}"
            sheet.cell(row=row, column=1, value=method_label)
            sheet.cell(row=row, column=2, value=SCENARIO_LABELS[scenario])
            _formula(sheet.cell(row=row, column=3), f"={assumption_ref}", assumption_format)
            if method_key == "gordon_growth":
                terminal_formula = f'=IF({WACC_CELL}<={assumption_ref},"",{terminal_metric_ref}*(1+{assumption_ref})/({WACC_CELL}-{assumption_ref}))'
            else:
                terminal_formula = f"={terminal_metric_ref}*{assumption_ref}"
            _formula(sheet.cell(row=row, column=4), terminal_formula, CURRENCY_FORMAT)
            _formula(
                sheet.cell(row=row, column=5),
                f'=IF(D{row}="","",D{row}/(1+{WACC_CELL})^{PROJECTION_YEARS_CELL})',
                CURRENCY_FORMAT,
            )
            _formula(sheet.cell(row=row, column=6), f'=IF(E{row}="","",$B$6+E{row})', CURRENCY_FORMAT)
            _formula(sheet.cell(row=row, column=7), "=$B$5", CURRENCY_FORMAT)
            _formula(sheet.cell(row=row, column=8), f'=IF(F{row}="","",F{row}-G{row})', CURRENCY_FORMAT)
            _formula(sheet.cell(row=row, column=9), f'=IF(OR(H{row}="",$B$4<=0),"",H{row}/$B$4)', PRICE_FORMAT)
            _formula(sheet.cell(row=row, column=10), f'=IF(OR(I{row}="",$B$3<=0),"",I{row}/$B$3-1)', PERCENT_FORMAT)
            row_map[(method_key, scenario)] = row
            row += 1

    _subtotal_border(sheet, 11, 10)
    _subtotal_border(sheet, 14, 10)
    _subtotal_border(sheet, 17, 10)
    _set_widths(
        sheet,
        {"A": 20, "B": 12, "C": 18, "D": 18, "E": 18, "F": 18, "G": 16, "H": 18, "I": 14, "J": 18},
    )
    sheet.freeze_panes = "C9"
    return row_map


def _build_summary_sheet(
    sheet: Worksheet, valuation: Mapping[str, Any], valuation_rows: Mapping[tuple[str, str], int]
) -> None:
    _title(sheet, "DCF Model Summary", 7)
    _label_value_formula(sheet, 3, "Ticker", "=Assumptions!B3")
    _label_value_formula(sheet, 4, "Company Name", "=Assumptions!B4")
    _label_value_formula(sheet, 5, "Current Price", f"={CURRENT_PRICE_CELL}", PRICE_FORMAT)
    _label_value_formula(sheet, 6, "Base Revenue", f"={BASE_REVENUE_CELL}", CURRENCY_FORMAT)
    _label_value_formula(sheet, 7, "Projection Years", f"={PROJECTION_YEARS_CELL}", NUMBER_FORMAT)
    _label_value_formula(sheet, 8, "Model Status", "=Checks!B6")

    _section(sheet, 10, "Base Case Valuation")
    _header_row(sheet, 11, ["Method", "Value/Share", "Upside/(Downside)", "Enterprise Value", "Equity Value"])
    summary_methods = [
        ("Gordon Growth", "gordon_growth"),
        ("EV/EBITDA Exit", "ev_ebitda_exit"),
        ("EV/Revenue Exit", "ev_revenue_exit"),
    ]
    for idx, (label, key) in enumerate(summary_methods, start=12):
        valuation_row = valuation_rows[(key, "base")]
        sheet.cell(row=idx, column=1, value=label)
        _formula(sheet.cell(row=idx, column=2), f"=Valuation!I{valuation_row}", PRICE_FORMAT)
        _formula(sheet.cell(row=idx, column=3), f"=Valuation!J{valuation_row}", PERCENT_FORMAT)
        _formula(sheet.cell(row=idx, column=4), f"=Valuation!F{valuation_row}", CURRENCY_FORMAT)
        _formula(sheet.cell(row=idx, column=5), f"=Valuation!H{valuation_row}", CURRENCY_FORMAT)

    sheet["A16"] = "Workbook recalculates when assumption cells on the Assumptions sheet are changed."
    sheet["A16"].font = NOTE_FONT
    _set_widths(sheet, {"A": 24, "B": 16, "C": 18, "D": 18, "E": 18, "F": 14, "G": 14})


def _build_historical_sheet(sheet: Worksheet, historical: Mapping[str, Any]) -> None:
    _title(sheet, "Historical Support", 8)
    row = 3
    sections = [
        (
            "EBITDA",
            "ebitda",
            [
                ("fiscal_year", "Fiscal Year"),
                ("revenue", "Revenue"),
                ("ebitda", "EBITDA"),
                ("ebitda_margin", "EBITDA Margin"),
                ("avg", "Average"),
            ],
            {"revenue", "ebitda"},
            {"ebitda_margin", "avg"},
            set(),
        ),
        (
            "Depreciation & Amortization",
            "depreciation",
            [
                ("fiscal_year", "Fiscal Year"),
                ("revenue", "Revenue"),
                ("da", "D&A"),
                ("da_pct_rev", "% Revenue"),
                ("avg", "Average"),
            ],
            {"revenue", "da"},
            {"da_pct_rev", "avg"},
            set(),
        ),
        (
            "Capital Expenditures",
            "capex",
            [
                ("fiscal_year", "Fiscal Year"),
                ("revenue", "Revenue"),
                ("capex", "CapEx"),
                ("capex_pct_rev", "% Revenue"),
                ("avg", "Average"),
            ],
            {"revenue", "capex"},
            {"capex_pct_rev", "avg"},
            set(),
        ),
        (
            "Net Working Capital",
            "nwc",
            [
                ("fiscal_year", "Fiscal Year"),
                ("revenue", "Revenue"),
                ("nwc", "NWC"),
                ("nwc_pct_rev", "% Revenue"),
                ("avg", "Average"),
            ],
            {"revenue", "nwc"},
            {"nwc_pct_rev", "avg"},
            set(),
        ),
        (
            "EV / EBITDA Multiple",
            "ev_ebitda",
            [
                ("quarter_end", "Quarter End"),
                ("ev", "Enterprise Value"),
                ("ev_ebitda", "EV/EBITDA"),
                ("avg", "Average"),
            ],
            {"ev"},
            set(),
            {"ev_ebitda", "avg"},
        ),
        (
            "EV / Revenue Multiple",
            "rev_multiple",
            [
                ("quarter_end", "Quarter End"),
                ("ev", "Enterprise Value"),
                ("ev_revenue", "EV/Revenue"),
                ("avg", "Average"),
            ],
            {"ev"},
            set(),
            {"ev_revenue", "avg"},
        ),
    ]
    for title, key, columns, currency_fields, percent_fields, multiple_fields in sections:
        row = _write_historical_section(
            sheet,
            row,
            title,
            _sequence_of_mappings(historical.get(key)),
            columns,
            currency_fields,
            percent_fields,
            multiple_fields,
        )
        row += 2

    _set_widths(sheet, {"A": 18, "B": 18, "C": 18, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16})
    sheet.freeze_panes = "A3"


def _build_checks_sheet(sheet: Worksheet, projection_years: int) -> None:
    _title(sheet, "Model Checks", 6)
    _header_row(sheet, 2, ["Check", "Status", "Notes"])
    rows = [
        (
            "Source inputs present",
            f'=IF(AND({SHARES_CELL}>0,{BASE_REVENUE_CELL}>0),"OK","Check")',
            "Shares and base revenue are required.",
        ),
        (
            "Terminal growth below WACC",
            f'=IF(AND({WACC_CELL}>Assumptions!B27,{WACC_CELL}>Assumptions!C27,{WACC_CELL}>Assumptions!D27),"OK","Check")',
            "Required for Gordon Growth scenarios.",
        ),
        (
            "Projection length",
            f'=IF(AND({PROJECTION_YEARS_CELL}>=5,{PROJECTION_YEARS_CELL}<=8,{PROJECTION_YEARS_CELL}={projection_years}),"OK","Check")',
            "Supported export range is 5-8 years.",
        ),
    ]
    for idx, (label, formula, note) in enumerate(rows, start=3):
        sheet.cell(row=idx, column=1, value=label)
        _formula(sheet.cell(row=idx, column=2), formula)
        sheet.cell(row=idx, column=3, value=note)
    sheet.cell(row=6, column=1, value="Overall Status").font = LABEL_FONT
    _formula(sheet.cell(row=6, column=2), '=IF(COUNTIF(B3:B5,"Check")=0,"OK","Check")')
    sheet["B6"].font = Font(bold=True, color="000000")
    for row in range(3, 7):
        sheet.cell(row=row, column=2).fill = OK_FILL
    _set_widths(sheet, {"A": 28, "B": 14, "C": 52})


def _build_sources_sheet(sheet: Worksheet, valuation: Mapping[str, Any], historical: Mapping[str, Any]) -> None:
    _title(sheet, "Sources & Audit", 6)
    _header_row(sheet, 2, ["Item", "Source", "As Of", "Notes"])
    generated = datetime.now(UTC).date()
    rows = [
        (
            "Company profile and market data",
            "yFinance",
            generated,
            "Current price, shares outstanding, debt, cash, WACC inputs, and fallback financials.",
        ),
        (
            "Quarterly multiples",
            str(historical.get("data_source") or "n/a"),
            generated,
            "SEC EDGAR is used when available; yFinance is the fallback.",
        ),
        ("User assumptions", "DCF page request", generated, "Editable input cells on the Assumptions sheet."),
        (
            "Valuation calculations",
            "Talisman DCF valuation service",
            generated,
            f"Ticker: {str(valuation.get('ticker') or '').upper()}",
        ),
    ]
    for idx, row_values in enumerate(rows, start=3):
        for col, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=idx, column=col, value=value)
            if col == 3:
                cell.number_format = DATE_FORMAT
    _set_widths(sheet, {"A": 30, "B": 28, "C": 14, "D": 74})
    for row in range(3, 3 + len(rows)):
        sheet.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")


def _scenario_row(sheet: Worksheet, row: int, label: str, values: Mapping[str, Any], number_format: str) -> None:
    _label(sheet.cell(row=row, column=1), label)
    for idx, scenario in enumerate(SCENARIOS, start=2):
        cell = sheet.cell(row=row, column=idx, value=_number_or_none(values.get(scenario)))
        cell.number_format = number_format
        cell.font = INPUT_FONT
        cell.fill = INPUT_FILL


def _write_historical_section(
    sheet: Worksheet,
    start_row: int,
    title: str,
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[tuple[str, str]],
    currency_fields: set[str],
    percent_fields: set[str],
    multiple_fields: set[str],
) -> int:
    _section(sheet, start_row, title)
    _header_row(sheet, start_row + 1, [label for _, label in columns])
    if not rows:
        sheet.cell(row=start_row + 2, column=1, value="No historical data available.").font = NOTE_FONT
        return start_row + 2

    for row_idx, item in enumerate(rows, start=start_row + 2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            value = item.get(key)
            cell = sheet.cell(row=row_idx, column=col_idx)
            if key in percent_fields:
                numeric = _number_or_none(value)
                cell.value = numeric / 100 if numeric is not None else None
                cell.number_format = PERCENT_FORMAT
            elif key in multiple_fields:
                cell.value = _number_or_none(value)
                cell.number_format = MULTIPLE_FORMAT
            elif key in currency_fields:
                cell.value = _number_or_none(value)
                cell.number_format = CURRENCY_FORMAT
            else:
                cell.value = value
            cell.border = THIN_BORDER
    return start_row + 1 + len(rows)


def _title(sheet: Worksheet, title: str, max_col: int) -> None:
    sheet.cell(row=1, column=1, value=title)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = sheet.cell(row=1, column=1)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24


def _section(sheet: Worksheet, row: int, title: str) -> None:
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER


def _header_row(sheet: Worksheet, row: int, values: Sequence[Any]) -> None:
    for col, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _label(cell, value: str) -> None:
    cell.value = value
    cell.font = LABEL_FONT
    cell.border = THIN_BORDER


def _formula(cell, formula: str, number_format: str = "General") -> None:
    cell.value = formula
    cell.number_format = number_format
    cell.font = FORMULA_FONT
    cell.border = THIN_BORDER


def _label_value_formula(sheet: Worksheet, row: int, label: str, formula: str, number_format: str = "General") -> None:
    _label(sheet.cell(row=row, column=1), label)
    _formula(sheet.cell(row=row, column=2), formula, number_format)


def _subtotal_border(sheet: Worksheet, row: int, max_col: int) -> None:
    border = Border(top=Side(style="thin", color="111827"))
    for col in range(1, max_col + 1):
        sheet.cell(row=row, column=col).border = border
        sheet.cell(row=row, column=col).font = Font(bold=True, color="000000")


def _set_widths(sheet: Worksheet, widths: Mapping[str, int]) -> None:
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def _series(value: Any, years: int) -> list[Any]:
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return list(value)[:years]
    return [value] * years


def _mapping(value: Any) -> Mapping[str, Any]:
    return value if isinstance(value, Mapping) else {}


def _sequence_of_mappings(value: Any) -> list[Mapping[str, Any]]:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes, bytearray)):
        return []
    return [item for item in value if isinstance(item, Mapping)]


def _number_or_none(value: Any) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None
